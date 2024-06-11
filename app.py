'''
____________________________________________________________________________________________

'''
# Libraries used for Retail Order Manager:

from openpyxl import Workbook, load_workbook
from datetime import datetime
import os.path
import pandas as pd


'''
____________________________________________________________________________________________

'''


print("--------Welcome to PyRetail Manager--------")

# Input file name with error handling
while True:
    file_path = input('\n'"Enter the file name you want to open or create (with .xlsx): ")
    if file_path.endswith(".xlsx"):
        break
    else:
        print("Invalid file name. Please include '.xlsx' extension.")

'''
____________________________________________________________________________________________
'''


# Global variables
workbook = None
sheet = None

# Function to load or create Excel workbook
def load_excel(file_path):
    global workbook, sheet
    try:
        if os.path.isfile(file_path):
            workbook = load_workbook(file_path)
            print(f"Opening {file_path}")
            sheet = workbook.active
        else:
            print("File not found. Creating a new Excel file.")
            create_excel()
            workbook = load_workbook(file_path)
            sheet = workbook.active
    except Exception as e:  # Catch any file-related errors
        print(f"An error occurred while loading or creating the file: {e}")
        return None, None  # Return None on error

    return workbook, sheet

# Function to create a new Excel file
def create_excel():
    global workbook, sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet 1"  # Add sheet title (optional)

    # Define column headers
    headers = ['Order ID', 'Product Name', 'Price', 'Quantity', 'Total Amount',
               'Customer Name', 'Phone Number', 'Email', 'Order Date & Time']
    sheet.append(headers)
    try:
        workbook.save(file_path)
        print(f"Excel file '{file_path}' created successfully.")
    except Exception as e:
        print(f"An error occurred while saving the file: {e}")
'''
____________________________________________________________________________________________

'''

# Function to add a new order
def add_order():
#    Adds a new order to the Excel file.
#    This function prompts the user to enter the details of the order, including the order ID, product name, price, quantity, customer name, phone number, and email. The function then calculates the total amount of the order by multiplying the price and quantity. The current date and time is also recorded.
#    The function appends the order details to the 'orders_3.xlsx' Excel file. The file is saved after adding the order.
#    This function does not take any parameters.
#    This function does not return any value.

    order_id = input("Enter Order ID: ")
    product_name = input("Enter Product Name: ")

    while True:
        try:
            price = float(input("Enter Price: "))
            if price <= 0:
                raise ValueError("Price must be a positive number.")
            break  # Exit the loop if input is valid
        except ValueError as e:
            print(f"Invalid input: {e}. Please enter a valid positive number.")

    while True:
        try:
            quantity = int(input("Enter Quantity: "))
            if quantity <= 0:
                raise ValueError("Quantity must be a positive integer.")
            break
        except ValueError as e:
            print(f"Invalid input: {e}. Please enter a valid positive integer.")

    total_amount = price * quantity
    customer_name = input("Enter Customer Name: ")
    phone_number = input("Enter Phone Number: ")
    email = input("Enter Email: ")
    date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # writing the data into excel file under the column heads
    sheet.append([order_id, product_name, price, quantity, total_amount, customer_name, phone_number, email, date_time])
    workbook.save(file_path)
    print("Order added successfully!"'\n')

'''
____________________________________________________________________________________________

'''

# Function to update an existing order
    # Updates an existing order based on the user's input for the new details.
    # Prompts the user to search by Order ID, then allows updating product name, price, quantity, customer name, phone number, and email.
    # Calculates the total amount based on the new price and quantity.
    # Updates the details in the Excel file 'orders_3.xlsx' and saves the workbook.
    # Prints success message if the order is updated, or an error message if the Order ID is not found.
     
def update_order():
    search_option = input("Enter 1 to search by Order ID or 2 to search by Customer Name: ")
    if search_option == '1':
        order_id = input("Enter the Order ID: ")
        found = False
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == order_id:
                found = True
                new_product_name = input("Enter the new product name (leave blank to keep the same): ") or row[1]
                
                while True:
                        try:
                            new_price = float(input("Enter the new price (leave blank to keep the same): ") or row[2])
                            if new_price <= 0:
                                raise ValueError("Price must be a positive number.")
                            break
                        except ValueError as e:
                            print(f"Invalid input: {e}. Please enter a valid positive number.")
                while True:
                        try:
                            new_quantity = int(input("Enter the new quantity (leave blank to keep the same): ") or row[3])
                            if new_quantity <= 0:
                                raise ValueError("Quantity must be a positive integer.")
                            break
                        except ValueError as e:
                            print(f"Invalid input: {e}. Please enter a valid positive integer.")

                new_total_amount = new_price * new_quantity
                new_customer_name = input("Enter the new customer name (leave blank to keep the same): ") or row[5]
                new_phone_number = input("Enter the new phone number (leave blank to keep the same): ") or row[6]
                new_email = input("Enter the new email (leave blank to keep the same): ") or row[7]
                new_date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Update values in the row
                row = list(row)  # Convert tuple to list to modify values
                row[1] = new_product_name
                row[2] = new_price
                row[3] = new_quantity
                row[4] = new_total_amount
                row[5] = new_customer_name
                row[6] = new_phone_number
                row[7] = new_email
                row[8] = new_date_time

                # Write updated row back to the Excel file
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=index, column=col_index, value=value)

                workbook.save(file_path)
                print("Order updated successfully!\n")
                break

        if not found:
            print("Order ID not found.")


'''
____________________________________________________________________________________________

'''

# function to convert data into tabular form to display data

from tabulate import tabulate
# Function to view order details
    # View order details based on search options.
    # This function prompts the user to select a search option: 1 for searching by Order ID or 2 for searching by Customer Name.
    # If the user selects 1, the function prompts the user to enter an Order ID and searches for the corresponding order details in the 'orders_3.xlsx' Excel file.
    # If the order is found, the function displays the order details in a tabular format using the 'tabulate' library.
    # If the order is not found, the function displays a message indicating that the Order ID was not found.
    # If the user selects 2, the function prompts the user to enter a Customer Name and searches for the corresponding order details in the 'orders_3.xlsx' Excel file.
    # If the customer is found, the function displays the order details in a tabular format using the 'tabulate' library.
    # If the customer is not found, the function displays a message indicating that the Customer Name was not found.
    # If the user enters an invalid option, the function displays a message indicating that the option is invalid.

def view_order_details():
    search_option = input("Enter 1 to search by Order ID or 2 to search by Customer Name: \n")
    if search_option == '1':
        order_id = input("\nEnter the Order ID: ")
        order_found = False
        order_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == order_id:
                order_found = True
                order_data.append(list(row))
        if order_found:
            columns = ['Order ID', 'Product Name', 'Price', 'Quantity', 'Total Amount', 'Customer Name', 'Phone Number', 'Email', 'Order Date & Time']
            order_df = pd.DataFrame(order_data, columns=columns)
            print(f"\nOrder Details for Order ID {order_id}:\n")
            print(tabulate(order_df, headers='keys', tablefmt='pretty'))
        else:
            print("Order ID not found.")

    elif search_option == '2':
        customer_name = input("\nEnter the Customer Name: ")
        customer_found = False
        customer_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[5].lower() == customer_name.lower():
                customer_found = True
                customer_data.append(list(row))
        if customer_found:
            columns = ['Order ID', 'Product Name', 'Price', 'Quantity', 'Total Amount', 'Customer Name', 'Phone Number', 'Email', 'Order Date & Time']
            customer_df = pd.DataFrame(customer_data, columns=columns)
            print(f"\nOrder Details for Customer '{customer_name}':\n")
            print(tabulate(customer_df, headers='keys', tablefmt='pretty'))
        else:
            print("Customer name not found.")
    else:
        print("Invalid option.")


'''
____________________________________________________________________________________________

'''
